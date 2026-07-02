import sqlite3
import os
import numpy as np
import pandas as pd
from scipy import stats
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series

class ExcelReporter:
    CANNIBALIZATION_THRESHOLD = 7.5
    COLOR_BOSCH = "E3170A"
    COLOR_SIEMENS = "00A19A"
    COLOR_HEADER_BG = "1F2937"
    COLOR_HEADER_FONT = "FFFFFF"
    COLOR_RISK = "FBE1E1"
    COLOR_RISK_FONT = "9C1F1F"
    COLOR_SAFE = "E3F3E8"
    COLOR_SAFE_FONT = "1E7A3D"
    COLOR_TITLE_BG = "111827"

    def __init__(self, db_name="data/bsh_fiyat_veritabani.db", report_name="reports/BSH_Kurumsal_BI_Raporu.xlsx", target_brands=("Bosch", "Siemens")):
        self.db_name = db_name
        self.report_name = report_name
        self.target_brands = list(target_brands)
        os.makedirs(os.path.dirname(self.report_name) or ".", exist_ok=True)

    def _detect_date_column(self, conn):
        cols = pd.read_sql_query("PRAGMA table_info(price_history)", conn)["name"].tolist()
        for candidate in ("scraped_at", "date", "created_at", "timestamp", "recorded_at", "updated_at"):
            if candidate in cols:
                return candidate
        return None

    def _fetch_data(self, conn):
        date_col = self._detect_date_column(conn)
        order_col = f"ph.{date_col}" if date_col else "ph.id"
        query = f"""
            SELECT p.brand         AS Marka,
                   p.category      AS Kategori,
                   p.model_code    AS Model_Kodu,
                   p.title         AS Urun_Adi,
                   ph.price        AS Fiyat,
                   ph.stock_status AS Stok_Durumu
            FROM products p
            JOIN price_history ph ON p.id = ph.product_id
            WHERE ph.id = (
                SELECT ph2.id FROM price_history ph2
                WHERE ph2.product_id = p.id
                ORDER BY {order_col.replace('ph.', 'ph2.')} DESC, ph2.id DESC
                LIMIT 1
            )
        """
        df = pd.read_sql_query(query, conn)
        df["Kategori"] = df["Kategori"].str.strip().str.title()
        return df

    def _build_cannibalization_table(self, df):
        df_bs = df[df["Marka"].isin(self.target_brands)]
        if df_bs.empty:
            return pd.DataFrame()
        
        stats_df = (
            df_bs.groupby(["Kategori", "Marka"])["Fiyat"]
            .agg(Ortalama_Fiyat="mean", Urun_Sayisi="count", Min_Fiyat="min", Max_Fiyat="max", Standart_Sapma="std")
            .reset_index()
        )
        wide = stats_df.pivot(index="Kategori", columns="Marka", values=["Ortalama_Fiyat", "Urun_Sayisi", "Min_Fiyat", "Max_Fiyat", "Standart_Sapma"])
        wide.columns = [f"{a}_{b}" for a, b in wide.columns]
        wide = wide.reset_index()
        
        b1, b2 = self.target_brands[0], self.target_brands[1]
        for brand in (b1, b2):
            if f"Ortalama_Fiyat_{brand}" not in wide.columns:
                wide[f"Ortalama_Fiyat_{brand}"] = np.nan
            if f"Urun_Sayisi_{brand}" not in wide.columns:
                wide[f"Urun_Sayisi_{brand}"] = 0
                
        wide = wide.dropna(subset=[f"Ortalama_Fiyat_{b1}", f"Ortalama_Fiyat_{b2}"]).copy()
        if wide.empty:
            return wide
            
        wide["Fiyat_Farki_TL"] = wide[f"Ortalama_Fiyat_{b2}"] - wide[f"Ortalama_Fiyat_{b1}"]
        wide[f"{b2}_Prim_Yuzde"] = (wide["Fiyat_Farki_TL"] / wide[f"Ortalama_Fiyat_{b1}"] * 100)
        
        p_values = []
        t_stats = []
        sig_results = []
        for cat in wide["Kategori"]:
            b1_prices = df_bs[(df_bs["Kategori"] == cat) & (df_bs["Marka"] == b1)]["Fiyat"].dropna()
            b2_prices = df_bs[(df_bs["Kategori"] == cat) & (df_bs["Marka"] == b2)]["Fiyat"].dropna()
            if len(b1_prices) > 1 and len(b2_prices) > 1:
                t_stat, p_val = stats.ttest_ind(b2_prices, b1_prices, equal_var=False)
                p_values.append(p_val)
                t_stats.append(t_stat)
                sig_results.append("Anlamlı" if p_val < 0.05 else "Tesadüfi")
            else:
                p_values.append(np.nan)
                t_stats.append(np.nan)
                sig_results.append("Veri Yetersiz")
                
        wide["P_Degeri"] = p_values
        wide["T_Istatistigi"] = t_stats
        wide["Istatistiksel_Fark"] = sig_results
        
        def label(x):
            if abs(x) < self.CANNIBALIZATION_THRESHOLD:
                return "Kanibalizm Riski"
            return f"{b2} Premium" if x > 0 else f"{b1} Premium"
            
        wide["Stratejik_Konum"] = wide[f"{b2}_Prim_Yuzde"].apply(label)
        return wide.sort_values(f"{b2}_Prim_Yuzde", ascending=False).reset_index(drop=True)

    def _generate_insight_text(self, wide):
        b1, b2 = self.target_brands[0], self.target_brands[1]
        if wide.empty:
            return "Ortak kategori bulunamadı, analiz yapılamıyor."
        overall_premium = wide[f"{b2}_Prim_Yuzde"].mean()
        risk_segments = wide[wide["Stratejik_Konum"] == "Kanibalizm Riski"]
        anlamli_segments = wide[wide["Istatistiksel_Fark"] == "Anlamlı"]
        yon = "daha pahalı" if overall_premium >= 0 else "daha ucuz"
        
        text = (f"PİYASA DİNAMİKLERİ VE JND (Just Noticeable Difference) İÇGÖRÜSÜ: İstatistiksel Welch's T-Testi sonuçlarına göre; pazar genelinde {b2} ürünleri, "
                f"aynı segmentteki {b1} ürünlerine kıyasla ortalama %{abs(overall_premium):.1f} {yon} konumlandırılmıştır. Tüketici psikolojisinde eşik değer olan %{self.CANNIBALIZATION_THRESHOLD} baz alınmıştır. ")
        if not anlamli_segments.empty:
            text += f"Bu fiyat makası, analiz edilen {len(anlamli_segments)} kategoride (p<0.05 güven aralığında) istatistiksel olarak 'Anlamlı' bulunmuş olup, BSH grubunun bilinçli marka ayrıştırma stratejisini doğrulamaktadır. "
        if not risk_segments.empty:
            risky_names = ", ".join(risk_segments["Kategori"].tolist())
            text += (f"DİKKAT: {risky_names} segmentlerinde markalar arası fiyat farkı hedeflenen "
                     f"%{self.CANNIBALIZATION_THRESHOLD:.1f} (JND) barajının altında kalmıştır. Çapraz fiyat esnekliğinin yüksek olduğu bu alanlarda cannibalization riski mevcuttur.")
        return text

    def _style_header_row(self, ws, row=1, n_cols=None):
        n_cols = n_cols or ws.max_column
        fill = PatternFill("solid", fgColor=self.COLOR_HEADER_BG)
        font = Font(bold=True, color=self.COLOR_HEADER_FONT, size=11)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _autofit_columns(self, ws, min_width=15, max_width=50):
        widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col = cell.column_letter
                widths[col] = max(widths.get(col, 0), len(str(cell.value)))
        for col, width in widths.items():
            header_value = ws[f"{col}1"].value
            if header_value and "Fiyat" in str(header_value):
                ws.column_dimensions[col].width = 20
            elif header_value and "Standart_Sapma" in str(header_value):
                ws.column_dimensions[col].width = 20
            else:
                ws.column_dimensions[col].width = min(max(width + 4, min_width), max_width)

    def _write_title_block(self, ws, title, subtitle, span_cols=8):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(bold=True, size=16, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=self.COLOR_TITLE_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 30
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span_cols)
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = Font(italic=True, size=10, color="6B7280")
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18

    def _add_chart_comment(self, ws, text, start_col, start_row, end_col, end_row):
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=start_col, value=text)
        cell.font = Font(italic=True, color="1F2937", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")
        
        thick_border = Side(style='medium', color="D1D5DB")
        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].height = 20
            for c in range(start_col, end_col + 1):
                border = Border(
                    left=thick_border if c == start_col else None,
                    right=thick_border if c == end_col else None,
                    top=thick_border if r == start_row else None,
                    bottom=thick_border if r == end_row else None
                )
                ws.cell(row=r, column=c).border = border

    def generate_report(self):
        conn = sqlite3.connect(self.db_name)
        df = self._fetch_data(conn)
        conn.close()
        
        if df.empty:
            print("Veritabanında raporlanacak veri bulunamadı.")
            return
            
        b1, b2 = self.target_brands[0], self.target_brands[1]
        wide = self._build_cannibalization_table(df)
        insight_text = self._generate_insight_text(wide)
        
        with pd.ExcelWriter(self.report_name, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, sheet_name="Yonetici_Ozeti", index=False)
            wide.to_excel(writer, sheet_name="Kanibalizm_Analizi", index=False, startrow=0)
            df.sort_values(["Kategori", "Marka", "Fiyat"]).to_excel(writer, sheet_name="Detayli_Liste", index=False)
            
        wb = load_workbook(self.report_name)
        
        # --- DETAYLI LİSTE ---
        ws_det = wb["Detayli_Liste"]
        self._style_header_row(ws_det)
        self._autofit_columns(ws_det, min_width=15, max_width=60)
        ws_det.freeze_panes = "A2"
        for row in ws_det.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = '#,##0.00 "TL"'

        # --- KANİBALİZM ANALİZİ & VARYANS ---
        ws_kan = wb["Kanibalizm_Analizi"]
        self._style_header_row(ws_kan)
        ws_kan.freeze_panes = "A2"
        headers = [c.value for c in ws_kan[1]]
        pct_col = headers.index(f"{b2}_Prim_Yuzde") + 1 if f"{b2}_Prim_Yuzde" in headers else None
        konum_col = headers.index("Stratejik_Konum") + 1 if "Stratejik_Konum" in headers else None
        ist_col = headers.index("Istatistiksel_Fark") + 1 if "Istatistiksel_Fark" in headers else None
        
        price_cols = [i + 1 for i, h in enumerate(headers) if h and ("Fiyat" in h or "Standart_Sapma" in h)]
        for col in price_cols:
            for row in ws_kan.iter_rows(min_row=2, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '#,##0.00 "TL"'
                    
        if pct_col:
            for row in ws_kan.iter_rows(min_row=2, min_col=pct_col, max_col=pct_col):
                for cell in row:
                    cell.number_format = '+0.0"%";-0.0"%"'
                    
        if konum_col:
            risk_fill = PatternFill("solid", fgColor=self.COLOR_RISK)
            risk_font = Font(color=self.COLOR_RISK_FONT, bold=True)
            safe_fill = PatternFill("solid", fgColor=self.COLOR_SAFE)
            safe_font = Font(color=self.COLOR_SAFE_FONT)
            for r in range(2, ws_kan.max_row + 1):
                cell = ws_kan.cell(row=r, column=konum_col)
                target_fill = risk_fill if cell.value == "Kanibalizm Riski" else safe_fill
                target_font = risk_font if cell.value == "Kanibalizm Riski" else safe_font
                for c in range(1, ws_kan.max_column + 1):
                    ws_kan.cell(row=r, column=c).fill = target_fill
                ws_kan.cell(row=r, column=konum_col).font = target_font
                ist_cell = ws_kan.cell(row=r, column=ist_col)
                if ist_cell.value == "Anlamlı":
                    ist_cell.font = Font(color="005CBF", bold=True)
                elif ist_cell.value == "Tesadüfi":
                    ist_cell.font = Font(color="856404")
                    
        self._autofit_columns(ws_kan, min_width=16, max_width=35)

        k_row = ws_kan.max_row + 3
        ws_kan.merge_cells(start_row=k_row, start_column=1, end_row=k_row, end_column=10)
        ws_kan.cell(row=k_row, column=1, value="🔬 Gelişmiş İstatistiksel Analiz ve Varyans Metodolojisi").font = Font(bold=True, size=14, color="FFFFFF")
        ws_kan.cell(row=k_row, column=1).fill = PatternFill("solid", fgColor="111827")
        ws_kan.row_dimensions[k_row].height = 25

        text_meth = ("Bu sekmedeki veriler, iki marka arasındaki fiyat farklarının tesadüfi veri dalgalanması mı yoksa şirketin bilinçli bir pazar stratejisi mi olduğunu kanıtlamak için Welch's T-Testi ile analiz edilmiştir.\n\n"
                     f"• Eşik Değeri (JND): Tüketici psikolojisi baz alınarak çapraz fiyat esnekliği riski (Kanibalizm) için algılanabilir minimum fark %{self.CANNIBALIZATION_THRESHOLD} olarak belirlenmiştir.\n"
                     "• P-Değeri (P-Value) < 0.05: İlgili kategorideki fiyat makasının %95 istatistiksel güven seviyesiyle anlamlı olduğunu ve bilinçli bir konumlandırma yapıldığını kanıtlar.\n"
                     "• Standart Sapma (σ): Markanın ilgili kategorideki fiyat esnekliğini gösterir. Yüksek standart sapma değeri, markanın o segmentte giriş seviyesinden premium seviyeye kadar geniş bir fiyat/ürün yelpazesi sunduğunu işaret eder.")
        ws_kan.merge_cells(start_row=k_row+1, start_column=1, end_row=k_row+6, end_column=10)
        cell_meth = ws_kan.cell(row=k_row+1, column=1, value=text_meth)
        cell_meth.font = Font(size=11, color="1F2937")
        cell_meth.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        cell_meth.fill = PatternFill("solid", fgColor="F9FAFB")
        for r in range(k_row+1, k_row+7):
            ws_kan.row_dimensions[r].height = 18

        v_row = k_row + 8
        ws_kan.cell(row=v_row, column=1, value="Pazar Esnekliği ve Varyans Detayları").font = Font(bold=True, size=13)
        var_headers = ["Kategori", f"{b1} Ürün Adedi", f"{b2} Ürün Adedi", f"{b1} Std Sapma (σ)", f"{b2} Std Sapma (σ)", "T-İstatistiği (Değeri)"]
        for i, h in enumerate(var_headers):
            c = ws_kan.cell(row=v_row+1, column=1+i, value=h)
            c.font = Font(bold=True, color=self.COLOR_HEADER_FONT)
            c.fill = PatternFill("solid", fgColor=self.COLOR_HEADER_BG)
        
        for i, r in wide.iterrows():
            row_idx = v_row + 2 + i
            ws_kan.cell(row=row_idx, column=1, value=r["Kategori"]).font = Font(bold=True)
            ws_kan.cell(row=row_idx, column=2, value=r[f"Urun_Sayisi_{b1}"])
            ws_kan.cell(row=row_idx, column=3, value=r[f"Urun_Sayisi_{b2}"])
            v1 = ws_kan.cell(row=row_idx, column=4, value=r[f"Standart_Sapma_{b1}"] if not pd.isna(r[f"Standart_Sapma_{b1}"]) else 0)
            v2 = ws_kan.cell(row=row_idx, column=5, value=r[f"Standart_Sapma_{b2}"] if not pd.isna(r[f"Standart_Sapma_{b2}"]) else 0)
            t_val = ws_kan.cell(row=row_idx, column=6, value=r["T_Istatistigi"] if not pd.isna(r["T_Istatistigi"]) else 0)
            v1.number_format = '#,##0.00 "TL"'
            v2.number_format = '#,##0.00 "TL"'
            t_val.number_format = '0.000'

        # --- YÖNETİCİ ÖZETİ (DASHBOARD KOKPİTİ) ---
        ws_sum = wb["Yonetici_Ozeti"]
        self._write_title_block(ws_sum, "BSH Kurumsal İş Zekası (BI) ve Strateji Kokpiti", f"{b1} vs {b2} | Executive Dashboard | Oluşturulma: {pd.Timestamp.now().strftime('%d.%m.%Y %H:%M')}", span_cols=8)
        
        ws_sum.merge_cells(start_row=4, start_column=1, end_row=7, end_column=8)
        insight_cell = ws_sum.cell(row=4, column=1, value=insight_text)
        insight_cell.font = Font(size=11, bold=True, color="1F2937")
        insight_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        insight_cell.fill = PatternFill("solid", fgColor="E5E7EB")
        for r in range(4, 8):
            ws_sum.row_dimensions[r].height = 22

        kpi_row = 9
        ws_sum.cell(row=kpi_row, column=1, value="Makro İstatistikler").font = Font(bold=True, size=13)
        kpi_headers = ["KPI Göstergesi", "Değer"]
        for i, h in enumerate(kpi_headers):
            c = ws_sum.cell(row=kpi_row + 1, column=1 + i, value=h)
            c.font = Font(bold=True, color=self.COLOR_HEADER_FONT)
            c.fill = PatternFill("solid", fgColor="374151")

        n_total_categories = wide.shape[0]
        n_risk = int((wide["Stratejik_Konum"] == "Kanibalizm Riski").sum()) if not wide.empty else 0
        n_sig = int((wide["Istatistiksel_Fark"] == "Anlamlı").sum()) if not wide.empty else 0
        avg_premium = wide[f"{b2}_Prim_Yuzde"].mean() if not wide.empty else np.nan

        kpis = [
            ("Analiz Edilen Ortak Segment", n_total_categories),
            ("İstatistiksel Anlamlı Hedefleme", n_sig),
            ("Kanibalizm Riski Taşıyan Segment", n_risk),
            (f"Genel Ortalama {b2} Primi (%)", None if pd.isna(avg_premium) else round(avg_premium, 1)),
            (f"Sistemdeki Toplam {b1} Ürünü", int((df["Marka"] == b1).sum())),
            (f"Sistemdeki Toplam {b2} Ürünü", int((df["Marka"] == b2).sum())),
        ]
        for i, (label, value) in enumerate(kpis, start=1):
            ws_sum.cell(row=kpi_row + 1 + i, column=1, value=label).font = Font(bold=True)
            ws_sum.cell(row=kpi_row + 1 + i, column=2, value=value)

        detail_row = kpi_row + len(kpis) + 3
        ws_sum.cell(row=detail_row, column=1, value="Veri Küpü (Data Cube)").font = Font(bold=True, size=13)
        mini_headers = ["Kategori", f"{b1} Ürün", f"{b2} Ürün", f"{b1} Ort. (TL)", f"{b2} Ort. (TL)", "Prim (%)", "P-Value", "Konum"]
        for i, h in enumerate(mini_headers):
            c = ws_sum.cell(row=detail_row + 1, column=1 + i, value=h)
            c.font = Font(bold=True, color=self.COLOR_HEADER_FONT)
            c.fill = PatternFill("solid", fgColor="374151")

        for i, r in wide.iterrows():
            row_idx = detail_row + 2 + i
            ws_sum.cell(row=row_idx, column=1, value=r["Kategori"]).font = Font(bold=True)
            ws_sum.cell(row=row_idx, column=2, value=r[f"Urun_Sayisi_{b1}"])
            ws_sum.cell(row=row_idx, column=3, value=r[f"Urun_Sayisi_{b2}"])
            ws_sum.cell(row=row_idx, column=4, value=round(r[f"Ortalama_Fiyat_{b1}"], 2)).number_format = '#,##0 "TL"'
            ws_sum.cell(row=row_idx, column=5, value=round(r[f"Ortalama_Fiyat_{b2}"], 2)).number_format = '#,##0 "TL"'
            ws_sum.cell(row=row_idx, column=6, value=round(r[f"{b2}_Prim_Yuzde"], 1)).number_format = '+0.0"%";-0.0"%"'
            ws_sum.cell(row=row_idx, column=7, value=r["P_Degeri"]).number_format = '0.000'
            
            konum_cell = ws_sum.cell(row=row_idx, column=8, value=r["Stratejik_Konum"])
            if r["Stratejik_Konum"] == "Kanibalizm Riski":
                fill, font = PatternFill("solid", fgColor=self.COLOR_RISK), Font(color=self.COLOR_RISK_FONT, bold=True)
            else:
                fill, font = PatternFill("solid", fgColor=self.COLOR_SAFE), Font(color=self.COLOR_SAFE_FONT)
            for c in range(1, 9):
                ws_sum.cell(row=row_idx, column=c).fill = fill
            konum_cell.font = font

        self._autofit_columns(ws_sum, min_width=13, max_width=35)
        ws_sum.sheet_view.showGridLines = False

        # --- DASHBOARD GRAFİKLERİ VE YORUMLAR (Genişletilmiş Layout) ---
        if not wide.empty:
            cat_ref = Reference(ws_sum, min_col=1, min_row=detail_row + 2, max_row=detail_row + 1 + len(wide))

            # Grafik 1: Fiyat Konumlandırması
            ch1 = BarChart()
            ch1.type = "col"
            ch1.style = 11
            ch1.title = "Kategori Bazlı Fiyat Konumlandırması"
            ch1.y_axis.title = "Ortalama Fiyat (TL)"
            data1 = Reference(ws_sum, min_col=4, max_col=5, min_row=detail_row + 1, max_row=detail_row + 1 + len(wide))
            ch1.add_data(data1, titles_from_data=True)
            ch1.set_categories(cat_ref)
            ch1.width, ch1.height = 15, 9
            ws_sum.add_chart(ch1, "J9")
            self._add_chart_comment(ws_sum, f"💡 Yönetici Özeti: {b1} ve {b2} arasındaki nominal fiyat makasını gösterir. Barların birbirine en çok yaklaştığı kategoriler rekabetin şirket içinde en yoğun yaşandığı, ürün yamyamlığı riskinin en yüksek olduğu alanlardır.", 10, 27, 18, 29)

            # Grafik 2: Marka Primi
            ch2 = LineChart()
            ch2.style = 13
            ch2.title = f"{b2} Fiyat Primi Dağılımı (%)"
            ch2.y_axis.title = "Prim Yüzdesi (%)"
            data2 = Reference(ws_sum, min_col=6, max_col=6, min_row=detail_row + 1, max_row=detail_row + 1 + len(wide))
            ch2.add_data(data2, titles_from_data=True)
            ch2.set_categories(cat_ref)
            ch2.width, ch2.height = 15, 9
            ws_sum.add_chart(ch2, "U9")
            self._add_chart_comment(ws_sum, f"💡 Yönetici Özeti: Tüketici psikolojisindeki JND (Just Noticeable Difference) limiti olan %{self.CANNIBALIZATION_THRESHOLD} sınırına göre markanın oransal primi. Çizginin bu eşiğin altına düşmesi durumunda alt markanın üst markayı satışlarda baskılaması beklenir.", 21, 27, 29, 29)

            # Grafik 3: Portföy Derinliği
            ch3 = BarChart()
            ch3.type = "bar"
            ch3.style = 10
            ch3.title = "Portföy Derinliği (Ürün Adedi)"
            ch3.x_axis.title = "Adet"
            data3 = Reference(ws_sum, min_col=2, max_col=3, min_row=detail_row + 1, max_row=detail_row + 1 + len(wide))
            ch3.add_data(data3, titles_from_data=True)
            ch3.set_categories(cat_ref)
            ch3.width, ch3.height = 15, 9
            ws_sum.add_chart(ch3, "J32")
            self._add_chart_comment(ws_sum, "💡 Yönetici Özeti: Her bir markanın ilgili kategoride canlı yayında tuttuğu benzersiz model (SKU) sayısını gösterir. Düşük ürün sayısına sahip marka, o segmentte pazar payını rakibe veya grup içi diğer markaya kaptırma riski taşır.", 10, 50, 18, 52)

            # Grafik 4: İstatistiksel Anlamlılık (P-Value)
            ch4 = LineChart()
            ch4.style = 12
            ch4.title = "T-Testi P-Value Anlamlılık Eşiği"
            ch4.y_axis.title = "P-Value"
            data4 = Reference(ws_sum, min_col=7, max_col=7, min_row=detail_row + 1, max_row=detail_row + 1 + len(wide))
            ch4.add_data(data4, titles_from_data=True)
            ch4.set_categories(cat_ref)
            ch4.width, ch4.height = 15, 9
            ws_sum.add_chart(ch4, "U32")
            self._add_chart_comment(ws_sum, "💡 Yönetici Özeti: İki marka arasındaki fiyatlandırma stratejisinin tesadüfi olup olmadığını ölçer. İstatistik bilimine göre P-Value değerinin 0.05'in altında olması, şirketin iki markayı son derece bilinçli ve kontrollü ayrıştırdığını matematiksel olarak kanıtlar.", 21, 50, 29, 52)

        wb.move_sheet("Yonetici_Ozeti", offset=-len(wb.sheetnames))
        wb.active = 0
        wb.save(self.report_name)
        print(f"Büyük BI Raporu başarıyla oluşturuldu: {self.report_name}")
        print("-" * 50)
        print(insight_text)

if __name__ == "__main__":
    reporter = ExcelReporter()
    reporter.generate_report()